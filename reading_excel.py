import openpyxl
import os

os.chdir('/home/stuart/Documents/automate_online_materials')
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
sheet = wb['Sheet1']
cell = sheet['A1']
print(sheet['A1'].value)
type(sheet['A1'].value)
print(sheet)
print(sheet.title)
print(sheet['C1'].value)
sheet.cell(row=1, column=2)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
