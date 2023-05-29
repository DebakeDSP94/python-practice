import openpyxl
import os

wb = openpyxl.workbook.Workbook()
wb.sheetnames
sheet = wb['Sheet']
sheet['A1'].value
os.chdir('/home/stuart/Documents/automate_online_materials')
wb.save('new_example_copy.xlsx')
sheet['A1'] = 'Hello world!'
sheet['A2'] = 42
wb.save('new_example_copy.xlsx')
sheet2 = wb.create_sheet()
wb.sheetnames
sheet2.title = 'My new sheet name'
wb.sheetnames
wb.save('new_example_copy.xlsx')
wb.create_sheet(index=0, title='My other sheet')
wb.sheetnames
wb.save('new_example_copy.xlsx')
